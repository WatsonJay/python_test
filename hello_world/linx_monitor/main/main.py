# -*- coding: utf-8 -*-
# @Author  : Jaywatson
# @File    : main.py
# @Software: PyCharm
import os
import random
import sys
import time

import openpyxl as openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from PyQt5.QtCore import pyqtSignal, QTimer
from PyQt5.QtWidgets import QApplication, QMainWindow, QDialog, QMessageBox, QWidget, QFileDialog
from PyQt5 import QtWidgets
import pyqtgraph as pg

from hello_world.linx_monitor.main.DownloadThread import DownloadThread
from hello_world.linx_monitor.main.config.config import config
from hello_world.linx_monitor.main.nmon_data_deal import nmon_data_deal
from hello_world.linx_monitor.main.threadClass import Thread
from hello_world.linx_monitor.main.totalDownload import TotalDownloadThread
from hello_world.linx_monitor.main.ui.Config_Dialog import Ui_Config_Dialog
from hello_world.linx_monitor.main.ui.MonitorWindow import Ui_Monitor_Window
from hello_world.linx_monitor.main.ui.analysis import Ui_nmonAnalysis_Form
from hello_world.linx_monitor.main.ui.config import Ui_SysConfig_Dialog
from hello_world.linx_monitor.main.ui.info import info_Form
from hello_world.linx_monitor.main.ui.simple import Ui_simple_Form
from hello_world.linx_monitor.main.ui.timer_Dialog import Ui_timer_Dialog
from hello_world.linx_monitor.main.ui.totalRun import Ui_totalRun_Form
from hello_world.linx_monitor.main.ui.totalanalysis import Ui_totanl_Form
from hello_world.linx_monitor.main.ui.user_helper import User_helper
from hello_world.linx_monitor.main.server_controller import Monitor_server
from hello_world.linx_monitor.main.ui.password_gui import Ui_password_Dialog

# 主页面
class MyMainWindow(QMainWindow, Ui_Monitor_Window):
    # 初始化
    def __init__(self, parent=None):
        super(MyMainWindow, self).__init__(parent)
        self.setupUi(self)
        self.loadConfig()
        self.actionshezhi.triggered.connect(self.sysConfigDialogShow)
        self.actioninfo.triggered.connect(self.infoShow)
        self.actionuse.triggered.connect(self.helperShow)
        self.actionnmon.triggered.connect(self.analysis_show)
        self.actionnmon_2.triggered.connect(self.totalRunShow)
        self.actiontoatalAnalysis.triggered.connect(self.totalAnalysisShow)
        self.serverAdd.clicked.connect(self.serverAdd_Open)
        self.serverModif.clicked.connect(self.serverModif_Open)
        self.serverDelete.clicked.connect(self.serverDel)
        self.server_listWidget.itemDoubleClicked.connect(self.showWidget)
        self.tabWidget.tabCloseRequested.connect(self.widget_Thread_close)

    # 读取系统配置
    def loadConfig(self):
        try:
            conf = config()
            SentionList = conf.getSection()
            self.server_listWidget.clear()
            self.server_listWidget.addItems(SentionList)
        except Exception as e:
            self.Tips("配置文件出现异常，请重置配置文件")

    # 打开介绍窗口
    def infoShow(self):
        self.infoShow = infoForm()
        self.infoShow.show()

    # 打开帮助窗口
    def helperShow(self):
        self.helperShow = userhelper()
        self.helperShow.show()

    # 打开批量运行窗口
    def totalRunShow(self):
        self.total_run_form = total_run_form()
        self.total_run_form.show()

    # 打开批量分析窗口
    def totalAnalysisShow(self):
        self.total_analysis_form = total_analysis_form()
        self.total_analysis_form.show()

    # 打开系统设置窗口
    def sysConfigDialogShow(self):
        try:
            conf = config()
            self.sysConfigDialog = sysConfigDialog()
            if self.sysConfigDialog.exec_():
                infos = self.sysConfigDialog.getDict()
                conf.cmdSection('sysconfig', infos)
            self.sysConfigDialog.destroy()
        except Exception as e:
            self.Tips("保存配置异常，请稍后重试")

    # 打开新增服务器窗口
    def serverAdd_Open(self):
        try:
            conf = config()
            self.serverConfigDialog = serverConfigDialog()
            if self.serverConfigDialog.exec_():
                infos = self.serverConfigDialog.getDict()
                conf.cmdSection(infos['name'], infos)
            self.serverConfigDialog.destroy()
            self.loadConfig()
        except Exception as e:
            self.Tips("保存配置异常，请稍后重试")

    # 打开编辑服务器窗口
    def serverModif_Open(self):
        try:
            conf = config()
            self.serverConfigDialog = serverConfigDialog()
            sention = self.server_listWidget.item(self.server_listWidget.currentRow()).text()
            info = conf.sentionAll(sention)
            self.serverConfigDialog.setDict(info)
            self.serverConfigDialog.disableName()
            if self.serverConfigDialog.exec_():
                infos = self.serverConfigDialog.getDict()
                conf.cmdSection(infos['name'], infos)
            self.serverConfigDialog.destroy()
            self.loadConfig()
        except:
            pass

    # 打开删除服务器窗口
    def serverDel(self):
        try:
            conf = config()
            sention = self.server_listWidget.item(self.server_listWidget.currentRow()).text()
            msg = QMessageBox.question(self, "删除警告", "你确定删除吗？", QMessageBox.Yes | QMessageBox.No,
                                       QMessageBox.No)  # 这里是固定格式，yes/no不能动
            # 判断消息的返回值
            if msg == QMessageBox.Yes:
                conf.delSection(sention)
            else:
                return
            self.loadConfig()
        except:
            pass

    # 打开服务器监控窗口
    def showWidget(self):
        try:
            conf = config()
            sention = self.server_listWidget.item(self.server_listWidget.currentRow()).text()
            info = conf.sentionAll(sention)
            ip = info['ip']
            self.serverConfigDialog = serverConfigDialog()
            self.serverConfigDialog.setDict(info)
            self.serverConfigDialog.disable()
            if self.serverConfigDialog.exec_():
                self.tab = simpleForm()
                self.tab.setIp(ip)
                self.tab.name = sention
                self.tab.checknmon()
                self.mointorThread = Thread()
                self.mointorThread.sendinfosSignal.connect(self.tab.set_data)
                self.mointorThread.sendExptionSignal.connect(self.tab.showMessage)
                self.tab.stop.connect(self.mointorThread.__del__)
                self.tab.nameSignal.connect(lambda: self.mointorThread.getInfos(sention))  # 通过信号槽设置名称
                self.tab.nameSignal.emit()
                self.tabWidget.addTab(self.tab, sention)
                self.tabWidget.setCurrentWidget(self.tab)
                self.mointorThread.start()
            self.serverConfigDialog.destroy()
        except Exception as e:
            self.Tips("线程启动异常，请重试或联系管理员")

    # 中止服务器监控页面线程窗口
    def widget_Thread_close(self, index):
        widget = self.tabWidget.widget(index)
        widget.stopThread()
        self.tabWidget.removeTab(index)

    # 打开分析窗口
    def analysis_show(self):
        self.analysisShow = analysis_form()
        self.analysisShow.show()
        self.analysisShow.loadfile('')

    # 提示窗口
    def Tips(self, message):
        QMessageBox.about(self, "提示", message)


# 服务器配置页面
class serverConfigDialog(QDialog, Ui_Config_Dialog):
    def __init__(self, parent=None):
        super(serverConfigDialog, self).__init__(parent)
        self.setupUi(self)

    def setDict(self, dict):
        conf = config()
        self.nameEdit.setText(dict['name'])
        self.ipEdit.setText(dict['ip'])
        self.portEdit.setText(dict['port'])
        self.userEdit.setText(dict['user'])
        self.passwordEdit.setText(conf.decrypt(dict['password']))

    def getDict(self):
        conf = config()
        info = {}
        info['name'] = self.nameEdit.text()
        info['ip'] = self.ipEdit.text()
        info['port'] = self.portEdit.text()
        info['user'] = self.userEdit.text()
        info['password'] = conf.encrypt(self.passwordEdit.text())
        return info

    def disableName(self):
        self.nameEdit.setDisabled(True)

    def disable(self):
        self.buttonBox.button(QtWidgets.QDialogButtonBox.Ok).setText('连接')
        self.nameEdit.setDisabled(True)
        self.ipEdit.setDisabled(True)
        self.portEdit.setDisabled(True)
        self.userEdit.setDisabled(True)
        self.passwordEdit.setDisabled(True)


# 性能监控统一页面
class simpleForm(QWidget, Ui_simple_Form):
    # 信号槽
    Downstop = pyqtSignal()
    stop = pyqtSignal()
    nameSignal = pyqtSignal()
    fileNameSignal = pyqtSignal()

    # 初始化
    def __init__(self):
        self.cpu_list = []
        self.mem_list = []
        self.disk_list = []
        self.nmon_infos = {}
        self.name = ''
        self.fileName = ''
        self.moveable = True
        super(simpleForm, self).__init__()
        self.setupUi(self)
        self.timer = QTimer(self)
        self.setWarnLine()
        self.visableUploadChange()
        self.visabledownloadChange()
        self.start_record.setDisabled(True)
        self.download_record.setDisabled(True)
        self.analysis_record.setDisabled(True)
        self.move_slot = pg.SignalProxy(self.graph_widget.scene().sigMouseMoved, rateLimit=60, slot=self.print_slot)
        self.upload_nmon.clicked.connect(self.uploadfile)
        self.moveable_btn.clicked.connect(self.moveType)
        self.start_record.clicked.connect(self.record_command)
        self.timer.timeout.connect(self.nmon_finished)
        self.download_record.clicked.connect(self.nmon_download)
        self.analysis_record.clicked.connect(self.analysis_show)

    # 设置预警线
    def setWarnLine(self):
        conf = config()
        self.CpuHLine.setPos(int(conf.getOption('sysconfig', 'cpuwarnline')))
        self.MemHLine.setPos(int(conf.getOption('sysconfig', 'memwarnline')))
        self.DiskHLine.setPos(int(conf.getOption('sysconfig', 'diskwarnline')))

    # 切换上传可见状态
    def visableUploadChange(self):
        self.upload_label.setHidden(bool(1 - self.upload_label.isHidden()))
        self.upload_progressBar.setHidden(bool(1 - self.upload_progressBar.isHidden()))

    # 切换下载可见状态
    def visabledownloadChange(self):
        self.download_label.setHidden(bool(1 - self.download_label.isHidden()))
        self.download_progressBar.setHidden(bool(1 - self.download_progressBar.isHidden()))

    # 填充ip
    def setIp(self, ip):
        self.ip_lineEdit.setText(ip)

    # 检测文件存在
    def checknmon(self):
        try:
            conf = config()
            infos = conf.sentionAll(self.name)
            monitor = Monitor_server()
            ssh = monitor.sshConnect(infos['ip'], int(infos['port']), infos['user'], conf.decrypt(infos['password']))
            nmon_checked = monitor.nmon_checked(ssh)
            if nmon_checked:
                self.nmon_label.setText('当前服务器已安装nmon，可进行性能监控')
                self.upload_nmon.setDisabled(True)
                self.start_record.setDisabled(False)
            else:
                self.nmon_label.setText('当前服务器未安装nmon，请点击上传')
            monitor.sshClose(ssh)
        except Exception as e:
            self.showMessage(str(e))
            pass

    # 上传文件
    def uploadfile(self):
        try:
            conf = config()
            infos = conf.sentionAll(self.name)
            monitor = Monitor_server()
            text = monitor.sftp_upload_file(infos['ip'], int(infos['port']), infos['user'],
                                            conf.decrypt(infos['password']))
            if text == '上传成功':
                self.showMessage('文件已上传')
                self.nmon_label.setText('当前服务器已安装nmon，可进行性能监控')
                self.upload_nmon.setDisabled(True)
                self.start_record.setDisabled(False)
            else:
                self.nmon_label.setText('当前服务器安装nmon失败，请点击再次上传')
        except Exception as e:
            self.showMessage(str(e))
            pass

    # 启动nmon记录
    def record_command(self):
        try:
            if self.start_record.text() == '开始':
                self.download_record.setDisabled(True)
                self.timerDialog = timerDialog()
                if self.timerDialog.exec_():
                    self.nmon_infos = self.timerDialog.getInfos()
                    if self.nmon_infos['name'] == '':
                        self.showMessage("名称未输入，无法启动")
                        return
                    conf = config()
                    infos = conf.sentionAll(self.name)
                    monitor = Monitor_server()
                    ssh = monitor.sshConnect(infos['ip'], int(infos['port']), infos['user'],
                                             conf.decrypt(infos['password']))
                    get_msgs = monitor.nmon_run(ssh, self.nmon_infos['name'], self.nmon_infos['time'], self.nmon_infos['tap'])
                    self.showMessage(get_msgs)
                    self.start_record.setText('中止')
                    self.record_name.setText(self.nmon_infos['name'] + '.nmon')
                    self.fileName = self.nmon_infos['name'] + '.nmon'
                    self.timer.start((int(self.nmon_infos['time']) * int(self.nmon_infos['tap']) + 1) * 1000)
                self.timerDialog.destroy()
            else:
                self.cancel()
        except Exception as e:
            self.showMessage(str(e))
            pass

    # 返回nmon记录
    def nmon_finished(self):
        self.start_record.setDisabled(False)
        self.download_record.setDisabled(False)
        self.showMessage('nmon运行结束')
        self.timer.stop()

    # 下载nmon记录
    def nmon_download(self):
        try:
            self.visabledownloadChange()
            self.DownloadThread = DownloadThread()
            self.nameSignal.connect(lambda: self.DownloadThread.getInfos(self.name))
            self.fileNameSignal.connect(lambda: self.DownloadThread.getfileName(self.fileName))
            self.Downstop.connect(self.DownloadThread.__del__)
            self.DownloadThread.sendDownExptionSignal.connect(self.showMessage)
            self.DownloadThread.sendSignal.connect(self.showProcess)
            self.nameSignal.emit()
            self.fileNameSignal.emit()
            self.DownloadThread.start()
        except Exception as e:
            self.showMessage(str(e))
            pass

    # 下载进度
    def showProcess(self, value):
        self.download_progressBar.setValue(value)
        if value == 100:
            self.visabledownloadChange()
            self.analysis_record.setDisabled(False)
            self.Downstop.emit()

    # 停止线程
    def stopThread(self):
        self.stop.emit()

    def cancel(self):
        try:
            conf = config()
            infos = conf.sentionAll(self.name)
            monitor = Monitor_server()
            ssh = monitor.sshConnect(infos['ip'], int(infos['port']), infos['user'], conf.decrypt(infos['password']))
            get_msg = monitor.nmon_cancel(ssh, self.nmon_infos['name'], self.nmon_infos['time'], self.nmon_infos['tap'])
            self.timer.stop()
            if get_msg == '取消成功':
                self.start_record.setText('开始')
                self.showMessage('已中止nmon运行')
        except Exception as e:
            self.showMessage(str(e))

    # 打开分析页
    def analysis_show(self):
        self.analysisShow = analysis_form()
        self.analysisShow.show()
        self.analysisShow.loadfile('temp/' + self.fileName)

    # 是否自动移动
    def moveType(self):
        if self.moveable == True:
            self.moveable_btn.setText('自动')
        else:
            self.moveable_btn.setText('锁定')
        self.moveable = bool(1 - self.moveable)

    # 绘制数据图形
    def set_data(self, cpu, mem, dick):
        try:
            if len(self.cpu_list) > 13 and self.moveable == True:
                self.graph_widget.setXRange(len(self.cpu_list) - 13, len(self.cpu_list) + 2, padding=0)
            self.showMessage('采样正常')
            self.cpu_list.append(cpu)
            self.mem_list.append(mem)
            self.disk_list.append(dick)
            self.cpu_line.setData(self.cpu_list)
            self.memory_line.setData(self.mem_list)
            self.disk_line.setData(self.disk_list)
        except Exception as e:
            pass

    # 响应鼠标移动绘制光标
    def print_slot(self, event=None):
        if event is None:
            print("事件为空")
        else:
            pos = event[0]  # 获取事件的鼠标位置
            try:
                # 如果鼠标位置在绘图部件中
                if self.graph_widget.sceneBoundingRect().contains(pos):
                    mousePoint = self.graph_widget.plotItem.vb.mapSceneToView(pos)  # 转换鼠标坐标
                    index = int(mousePoint.x())  # 鼠标所处的X轴坐标
                    if 0 <= index <= len(self.cpu_list) - 1:
                        # 在label中写入HTML
                        self.point_label.setHtml(
                            "<p style='color:white'>CPU：{0}%</p><p style='color:white'>内存：{1}%</p><p style='color:white'>磁盘：{2}%</p>".format(
                                self.cpu_list[index], self.mem_list[index], self.disk_list[index]))
                        self.point_label.setPos(mousePoint.x(), mousePoint.y())  # 设置label的位置
                        # 设置垂直线条和水平线条的位置组成十字光标
                        self.vLine.setPos(mousePoint.x())
            except Exception as e:
                pass

    # 显示日志
    def showMessage(self, msg):
        msg = "[" + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + "]" + msg
        self.info_borwser.append(msg)
        if "文件下载异常" in msg:
            self.visabledownloadChange()


# 系统设置页面
class sysConfigDialog(QDialog, Ui_SysConfig_Dialog):
    # 初始化
    def __init__(self, parent=None):
        super(sysConfigDialog, self).__init__(parent)
        self.setupUi(self)
        self.loadConfig()

    # 读取系统配置
    def loadConfig(self):
        try:
            conf = config()
            sysconfig = conf.sentionAll('sysconfig')
            self.CPU_spinBox.setValue(int(sysconfig['cpuwarnline']))
            self.MEM_spinBox.setValue(int(sysconfig['memwarnline']))
            self.DISK_spinBox.setValue(int(sysconfig['diskwarnline']))
            self.Simple_Rate_spinBox.setValue(int(sysconfig['simplerating']))
            self.password_lineEdit.setText(conf.decrypt(sysconfig['password']))
        except Exception as e:
            self.Tips("配置文件出现异常，请重置配置文件")

    # 保存系统配置
    def getDict(self):
        conf = config()
        info = {}
        info['cpuwarnline'] = str(self.CPU_spinBox.value())
        info['memwarnline'] = str(self.MEM_spinBox.value())
        info['diskwarnline'] = str(self.DISK_spinBox.value())
        info['simplerating'] = str(self.Simple_Rate_spinBox.value())
        info['password'] = conf.encrypt(self.password_lineEdit.text())
        return info

    # 提示窗口
    def Tips(self, message):
        QMessageBox.about(self, "提示", message)


# 介绍页面
class infoForm(QWidget, info_Form):
    # 初始化
    def __init__(self, parent=None):
        super(infoForm, self).__init__(parent)
        self.setupUi(self)


# 帮助页面
class userhelper(QWidget, User_helper):
    # 初始化
    def __init__(self, parent=None):
        super(userhelper, self).__init__(parent)
        self.setupUi(self)


# nmon启动参数页面
class timerDialog(QDialog, Ui_timer_Dialog):
    # 初始化
    def __init__(self, parent=None):
        super(timerDialog, self).__init__(parent)
        self.setupUi(self)

    def getInfos(self):
        infos = {}
        infos['time'] = str(self.time_spinBox.value())
        infos['tap'] = str(self.tap_spinBox.value())
        infos['name'] = self.fileName_lineEdit.text()
        return infos


# 分析页面
class analysis_form(QWidget, Ui_nmonAnalysis_Form):
    # 初始化
    def __init__(self, parent=None):
        super(analysis_form, self).__init__(parent)
        self.setupUi(self)
        self.prepare_list = locals()
        self.cwd = os.getcwd()
        self.cpu_user = []
        self.cpu_Idle = []
        self.mem = []
        self.disk_info = {}
        self.iops = []
        self.net_read = []
        self.net_write = []
        self.len = 0
        self.pushButton_2.clicked.connect(self.calculation_again)
        self.analysis_pushButton.clicked.connect(lambda: self.loadfile(''))
        self.pushButton_4.clicked.connect(self.write07Excel)
        self.moveStart_pushButton.clicked.connect(self.moveToStart)
        self.moveEnd_pushButton.clicked.connect(self.moveToEnd)
        self.reset_pushButton.clicked.connect(self.moveRestart)
        self.cpu_move_slot = pg.SignalProxy(self.cpu_widget.scene().sigMouseMoved, rateLimit=60,
                                            slot=self.print_cpu_slot)
        self.mem_move_slot = pg.SignalProxy(self.mem_widget.scene().sigMouseMoved, rateLimit=60,
                                            slot=self.print_mem_slot)
        self.iops_move_slot = pg.SignalProxy(self.iops_widget.scene().sigMouseMoved, rateLimit=60,
                                             slot=self.print_iops_slot)
        self.net_move_slot = pg.SignalProxy(self.net_widget.scene().sigMouseMoved, rateLimit=60,
                                            slot=self.print_net_slot)
        self.disk_move_slot = pg.SignalProxy(self.disk_widget.scene().sigMouseMoved, rateLimit=60,
                                             slot=self.print_disk_slot)

    # 加载数据
    def loadfile(self, filePath):
        try:
            if filePath == '':
                filePath, fileType = QFileDialog.getOpenFileName(self, "选取文件", "temp/", "nmon文件 (*.nmon);;所有文件 (*)")
            if filePath == '':
                return
            nmon = nmon_data_deal()
            self.textBrowser.clear()
            infos = nmon.file_read(filePath)
            self.analysis(infos)
            self.calculation(1, self.len)
            self.fileName_lineEdit.setText(filePath)
            self.simpleNum_label.setText(str(self.len) + '次')
            self.UI_init()
        except Exception as e:
            self.Tips('nmon文件解析异常，请重试')
            self.close()

    # 初始化处理
    def analysis(self, infos):
        for cpu in infos['cpu']:
            self.cpu_user.append(cpu[0] + cpu[1])
            self.cpu_Idle.append(cpu[2])
        for mem in infos['mem']:
            self.mem.append(round((mem[0] - mem[1] - mem[2] - mem[3]) / mem[0] * 100, 3))
        for net in infos['NET']:
            self.net_read.append(round(net[0] / 128, 3))
            self.net_write.append(round(net[1] / 128, 3))
        self.disk_info['diskName'] = infos['diskName']
        for diskName in infos['diskName']:
            self.disk_info[diskName + '_DISKBUSY'] = infos[diskName + '_DISKBUSY']
        self.iops = infos['DISKXFER']
        self.len = infos['simpleNumber']
        self.ip_lineEdit.setText(infos['ip'])
        self.infoShow('当前系统版本：' + infos['os'])

    def UI_init(self):
        self.star_spinBox.setMinimum(1)
        self.star_spinBox.setMaximum(self.len - 1)
        self.end_spinBox.setMinimum(2)
        self.end_spinBox.setMaximum(self.len)
        self.end_spinBox.setValue(self.len)
        self.cpu_user_line.setData(self.cpu_user)
        #self.cpu_idle_line.setData(self.cpu_Idle)
        self.mem_line.setData(self.mem)
        self.iops_line.setData(self.iops)
        self.net_read_line.setData(self.net_read)
        self.net_write_line.setData(self.net_write)
        for diskName in self.disk_info['diskName']:
            r = random.randint(0, 255)
            g = random.randint(0, 255)
            b = random.randint(0, 255)
            self.prepare_list[diskName + '_line'] = self.disk_widget.plot(pen=(r, g, b), name=diskName,
                                                                          symbolBrush=(r, g, b))
            self.prepare_list[diskName + '_line'].setData(self.disk_info[diskName + '_DISKBUSY'])

    def moveGraph(self, start, end):
        self.cpu_widget.setXRange(start, end, padding=0)  # 初始化X轴显示范围
        self.mem_widget.setXRange(start, end, padding=0)  # 初始化X轴显示范围
        self.disk_widget.setXRange(start, end, padding=0)  # 初始化X轴显示范围
        self.iops_widget.setXRange(start, end, padding=0)  # 初始化X轴显示范围
        self.net_widget.setXRange(start, end, padding=0)  # 初始化X轴显示范围

    def moveToStart(self):
        self.moveGraph(self.star_spinBox.value() - 5, self.star_spinBox.value() + 20)

    def moveToEnd(self):
        self.moveGraph(self.end_spinBox.value() - 5, self.end_spinBox.value() + 20)

    def moveRestart(self):
        self.moveGraph(0, 25)

    def calculation_again(self):
        self.calculation(self.star_spinBox.value(), self.end_spinBox.value())

    # 计算平均值
    def calculation(self, start, end):
        cpu_temp = 0
        mem_temp = 0
        net_temp = 0
        IOPS_temp = 0
        disk_temp = 0
        count = 0
        for i in range(start - 1, end):
            cpu_temp += self.cpu_user[i]
            mem_temp += self.mem[i]
            net_temp += self.net_read[i] + self.net_write[i]
            IOPS_temp += self.iops[i]
        for disk_name in self.disk_info['diskName']:
            temp1 = 0
            for i in range(start - 1, end):
                temp1 += self.disk_info[disk_name + '_DISKBUSY'][i]
            if round(temp1 / (end - start + 1), 2) != 0.0:
                disk_temp += temp1
                count += 1
        if count == 0:
            count = 1
        self.cpu_label.setText(str(round(cpu_temp / (end - start + 1), 2)) + '%')
        self.men_label.setText(str(round(mem_temp / (end - start + 1), 2)) + '%')
        self.disk_label.setText(str(round(disk_temp / count / (end - start + 1), 2)) + '%')
        self.net_label.setText(str(round(net_temp / (end - start + 1), 2)) + 'Mbps')
        self.IOPS_label.setText(str(round(IOPS_temp / (end - start + 1), 2)) + '次')

    # 导出
    def write07Excel(self):
        try:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = self.ip_lineEdit.text() + '统计表'
            value = [[self.label_7.text(), self.label_9.text(), self.label_13.text(), self.label_11.text(),
                      self.label_16.text()],
                     [self.cpu_label.text(), self.men_label.text(), self.disk_label.text(), self.net_label.text(),
                      self.IOPS_label.text()]]
            for i in range(0, 2):
                for j in range(0, len(value[i])):
                    sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
            filePath, ok2 = QFileDialog.getSaveFileName(self,
                                                        "文件保存",
                                                        "temp/",
                                                        "Xlsx Files (*.xlsx);;All Files (*)")
            if filePath == '':
                return
            wb.save(filePath)
            self.Tips('写入数据成功！')
        except Exception as e:
            self.Tips('导出execl异常，请重试')

    # 消息显示
    def infoShow(self, msg):
        self.textBrowser.append(msg)

    # 提示窗口
    def Tips(self, message):
        QMessageBox.about(self, "提示", message)

    # CPU光标
    def print_cpu_slot(self, event=None):
        if event is None:
            print("事件为空")
        else:
            pos = event[0]  # 获取事件的鼠标位置
            try:
                # 如果鼠标位置在绘图部件中
                if self.cpu_widget.sceneBoundingRect().contains(pos):
                    mousePoint = self.cpu_widget.plotItem.vb.mapSceneToView(pos)  # 转换鼠标坐标
                    index = int(mousePoint.x())  # 鼠标所处的X轴坐标
                    if 0 <= index < self.len:
                        # 在label中写入HTML
                        self.cpu_point_label.setHtml(
                            "<p style='color:yellow'>采样点：{0}</p><p style='color:white'>CPU使用率：{1}%</p><p style='color:white'>CPU空闲率：{2}%</p>".format(
                                str(index + 1), self.cpu_user[index], self.cpu_Idle[index]))
                        self.cpu_point_label.setPos(mousePoint.x(), mousePoint.y())  # 设置label的位置
                        # 设置垂直线条和水平线条的位置组成十字光标
                        self.cpu_vLine.setPos(mousePoint.x())
            except Exception as e:
                pass

    # mem光标
    def print_mem_slot(self, event=None):
        if event is None:
            print("事件为空")
        else:
            pos = event[0]  # 获取事件的鼠标位置
            try:
                # 如果鼠标位置在绘图部件中
                if self.mem_widget.sceneBoundingRect().contains(pos):
                    mousePoint = self.mem_widget.plotItem.vb.mapSceneToView(pos)  # 转换鼠标坐标
                    index = int(mousePoint.x())  # 鼠标所处的X轴坐标
                    if 0 <= index < len(self.mem):
                        # 在label中写入HTML
                        self.mem_point_label.setHtml(
                            "<p style='color:yellow'>采样点：{0}</p><p style='color:white'>内存用户使用率：{1}%</p>".format(
                                str(index + 1), self.mem[index]))
                        self.mem_point_label.setPos(mousePoint.x(), mousePoint.y())  # 设置label的位置
                        # 设置垂直线条和水平线条的位置组成十字光标
                        self.mem_vLine.setPos(mousePoint.x())
            except Exception as e:
                pass

    # disk光标
    def print_disk_slot(self, event=None):
        if event is None:
            print("事件为空")
        else:
            pos = event[0]  # 获取事件的鼠标位置
            try:
                # 如果鼠标位置在绘图部件中
                if self.disk_widget.sceneBoundingRect().contains(pos):
                    mousePoint = self.disk_widget.plotItem.vb.mapSceneToView(pos)  # 转换鼠标坐标
                    index = int(mousePoint.x())  # 鼠标所处的X轴坐标
                    html = ""
                    if 0 <= index < len(self.mem):
                        html = "<p style='color:yellow'>采样点：{}</p>".format(str(index + 1))
                        for diskName in self.disk_info['diskName']:
                            html += "<p style='color:white'>{0}使用率：{1}%</p>".format(diskName, self.disk_info[
                                diskName + '_DISKBUSY'][index])
                        # 在label中写入HTML
                        self.disk_point_label.setHtml(html)
                        self.disk_point_label.setPos(mousePoint.x(), mousePoint.y())  # 设置label的位置
                        # 设置垂直线条和水平线条的位置组成十字光标
                        self.disk_vLine.setPos(mousePoint.x())
            except Exception as e:
                pass

    # IOPS光标
    def print_iops_slot(self, event=None):
        if event is None:
            print("事件为空")
        else:
            pos = event[0]  # 获取事件的鼠标位置
            try:
                # 如果鼠标位置在绘图部件中
                if self.iops_widget.sceneBoundingRect().contains(pos):
                    mousePoint = self.iops_widget.plotItem.vb.mapSceneToView(pos)  # 转换鼠标坐标
                    index = int(mousePoint.x())  # 鼠标所处的X轴坐标
                    if 0 <= index < len(self.iops):
                        # 在label中写入HTML
                        self.iops_point_label.setHtml(
                            "<p style='color:yellow'>采样点：{0}</p><p style='color:white'>IO/sec：{1}次</p>".format(
                                str(index + 1), self.iops[index]))
                        self.iops_point_label.setPos(mousePoint.x(), mousePoint.y())  # 设置label的位置
                        # 设置垂直线条和水平线条的位置组成十字光标
                        self.iops_vLine.setPos(mousePoint.x())
            except Exception as e:
                pass

    # NET光标
    def print_net_slot(self, event=None):
        if event is None:
            print("事件为空")
        else:
            pos = event[0]  # 获取事件的鼠标位置
            try:
                # 如果鼠标位置在绘图部件中
                if self.net_widget.sceneBoundingRect().contains(pos):
                    mousePoint = self.net_widget.plotItem.vb.mapSceneToView(pos)  # 转换鼠标坐标
                    index = int(mousePoint.x())  # 鼠标所处的X轴坐标
                    if 0 <= index < len(self.net_read):
                        # 在label中写入HTML
                        self.net_point_label.setHtml(
                            "<p style='color:yellow'>采样点：{0}</p><p style='color:white'>net-读：{1}Mbps</p><p style='color:white'>net-写：{2}Mbps</p>".format(
                                str(index + 1), self.net_read[index], self.net_write[index]))
                        self.net_point_label.setPos(mousePoint.x(), mousePoint.y())  # 设置label的位置
                        # 设置垂直线条和水平线条的位置组成十字光标
                        self.net_vLine.setPos(mousePoint.x())
            except Exception as e:
                pass


# 批量运行
class total_run_form(QWidget, Ui_totalRun_Form):
    Downstop = pyqtSignal()
    stop = pyqtSignal()
    nameSignal = pyqtSignal()
    fileNameSignal = pyqtSignal()

    # 初始化
    def __init__(self, parent=None):
        super(total_run_form, self).__init__(parent)
        self.simplePages = locals()
        self.setupUi(self)
        self.loadConfig()
        self.count = 0
        self.download_pushButton.setDisabled(True)
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.nmon_finished)
        self.add_pushButton.clicked.connect(self.addserver)
        self.remove_pushButton.clicked.connect(self.deladded)
        self.start_pushButton.clicked.connect(self.totalrun)
        self.download_pushButton.clicked.connect(self.nmon_download)
        self.close_pushButton.clicked.connect(self.cancel)

    # 读取系统配置
    def loadConfig(self):
        try:
            conf = config()
            SentionList = conf.getSection()
            self.server_listWidget.clear()
            self.server_listWidget.addItems(SentionList)
        except Exception as e:
            self.Tips("配置文件出现异常，请重置配置文件")

    # 添加批处理服务器
    def addserver(self):
        text = self.server_listWidget.item(self.server_listWidget.currentRow()).text()
        exist = False
        for i in range(self.added_listWidget.count()):
            if text == self.added_listWidget.item(i).text():
                exist = True
        if exist == False:
            self.added_listWidget.addItem(text)

    # 删除批处理服务器
    def deladded(self):
        try:
            self.added_listWidget.takeItem(self.added_listWidget.currentRow())
        except:
            pass

    # 服务器批处理nmon
    def totalrun(self):
        try:
            conf = config()
            if self.added_listWidget.count() == 0:
                self.Tips("未添加服务器")
                return
            if self.groupName_lineEdit.text() == '':
                self.Tips("未输入组名")
                return
            self.remove_pushButton.setDisabled(True)
            self.add_pushButton.setDisabled(True)
            self.tap_spinBox.setDisabled(True)
            self.time_spinBox.setDisabled(True)
            self.download_pushButton.setDisabled(True)
            self.groupName_lineEdit.setDisabled(True)
            self.start_pushButton.setDisabled(True)
            self.time_progressBar.setValue(0)
            for i in range(self.added_listWidget.count()):
                sention = self.added_listWidget.item(i).text()
                infos = conf.sentionAll(sention)
                monitor = Monitor_server()
                ssh = monitor.sshConnect(infos['ip'], int(infos['port']), infos['user'],
                                         conf.decrypt(infos['password']))
                if monitor.nmon_checked(ssh) == False:
                    text = monitor.sftp_upload_file(infos['ip'], int(infos['port']), infos['user'],
                                                    conf.decrypt(infos['password']))
                get_msgs = monitor.nmon_run(ssh, self.groupName_lineEdit.text() + str(i),
                                            str(self.time_spinBox.value()),
                                            str(self.tap_spinBox.value()))
                if get_msgs != '启动运行,请稍后':
                    self.Tips('批量启动异常，请检查服务器配置，并重试')
                    self.remove_pushButton.setDisabled(False)
                    self.add_pushButton.setDisabled(False)
                    self.tap_spinBox.setDisabled(False)
                    self.time_spinBox.setDisabled(False)
                    self.groupName_lineEdit.setDisabled(False)
                    self.start_pushButton.setDisabled(False)
                monitor.sshClose(ssh)
            self.time_progressBar.setMaximum((self.time_spinBox.value() * self.tap_spinBox.value()))
            self.timer.start(1000)
            self.Tips('已启动nmon,请稍候')
            self.totalFile_label.setText(str(self.added_listWidget.count()))
        except Exception as e:
            self.Tips(str(e))
            self.remove_pushButton.setDisabled(False)
            self.add_pushButton.setDisabled(False)
            self.tap_spinBox.setDisabled(False)
            self.time_spinBox.setDisabled(False)
            self.groupName_lineEdit.setDisabled(False)
            self.start_pushButton.setDisabled(False)

    # 返回nmon记录
    def nmon_finished(self):
        if self.count == (self.time_spinBox.value() * self.tap_spinBox.value()):
            self.download_pushButton.setDisabled(False)
            self.Tips('nmon运行结束,可进行下载')
            self.timer.stop()
            self.count = 0
            self.time_progressBar.setValue(0)
            return
        self.count += 1
        self.time_progressBar.setValue((self.time_spinBox.value() * self.tap_spinBox.value())-self.count)

    # 下载nmon记录
    def nmon_download(self):
        try:
            serverList = []
            self.successFile_label.setText(str(0))
            self.failFile_label.setText(str(0))
            for i in range(self.added_listWidget.count()):
                serverList.append(self.added_listWidget.item(i).text())
            self.DownloadThread = TotalDownloadThread()
            self.nameSignal.connect(lambda: self.DownloadThread.getInfos(serverList))
            self.fileNameSignal.connect(lambda: self.DownloadThread.getfileName(self.groupName_lineEdit.text()))
            self.Downstop.connect(self.DownloadThread.__del__)
            self.DownloadThread.sendDownMsgSignal.connect(self.showDown)
            self.nameSignal.emit()
            self.fileNameSignal.emit()
            self.DownloadThread.start()
            self.start_pushButton.setDisabled(False)
        except Exception as e:
            self.Tips(str(e))

    # 下载文件数
    def showDown(self, text, value):
        if text == 'success':
            self.successFile_label.setText(str(int(self.successFile_label.text()) + 1))
        else:
            self.failFile_label.setText(str(int(self.failFile_label.text()) + 1))
        self.progressBar.setValue(value)
        if value == 100:
            self.Tips('下载完成')
            self.Downstop.emit()

    # 取消运行
    def cancel(self):
        try:
            conf = config()
            for i in range(self.added_listWidget.count()):
                sention = self.added_listWidget.item(i).text()
                infos = conf.sentionAll(sention)
                monitor = Monitor_server()
                ssh = monitor.sshConnect(infos['ip'], int(infos['port']), infos['user'],
                                         conf.decrypt(infos['password']))
                get_msgs = monitor.nmon_cancel(ssh, self.groupName_lineEdit.text() + str(i),
                                            str(self.time_spinBox.value()),
                                            str(self.tap_spinBox.value()))
                monitor.sshClose(ssh)
                self.timer.stop()
        except Exception as e:
            self.Tips(str(e))
        self.close()
    # 提示窗口
    def Tips(self, message):
        QMessageBox.about(self, "提示", message)


# 批量分析
class total_analysis_form(QWidget, Ui_totanl_Form):
    # 初始化
    def __init__(self, parent=None):
        super(total_analysis_form, self).__init__(parent)
        self.setupUi(self)
        self.analysisPage = locals()
        self.save_pushButton.setDisabled(True)
        self.save_pushButton.clicked.connect(self.totalsave)
        self.addFile_pushButton.clicked.connect(self.addFile)
        self.remove_pushButton.clicked.connect(self.delFile)
        self.analysis_pushButton.clicked.connect(self.totalanlysis)

    # 添加
    def addFile(self):
        filePaths, fileType = QFileDialog.getOpenFileNames(self, "选取文件", "temp/", "nmon文件 (*.nmon);;所有文件 (*)")
        if len(filePaths) == 0:
            return
        for filePath in filePaths:
            self.file_listWidget.addItem(filePath)
        self.save_pushButton.setDisabled(True)

    # 删除
    def delFile(self):
        try:
            self.file_listWidget.takeItem(self.file_listWidget.currentRow())
        except:
            pass

    # 分析
    def totalanlysis(self):
        try:
            if self.file_listWidget.count() == 0:
                self.Tips("未添加分析文件")
                return
            for page in self.analysisPage:
                if 'totalAnalysis' in page:
                    self.analysisPage[page].close()
            for i in range(self.file_listWidget.count()):
                file = self.file_listWidget.item(i).text()
                self.analysisPage['totalAnalysis' + str(i)] = analysis_form()
                self.analysisPage['totalAnalysis' + str(i)].show()
                self.analysisPage['totalAnalysis' + str(i)].loadfile(file)
            self.save_pushButton.setDisabled(False)
            self.analysis_pushButton.setDisabled(True)
        except:
            pass

    # 批量保存
    def totalsave(self):
        try:
            wb = openpyxl.Workbook()
            for i in range(self.file_listWidget.count()):
                self.analysisPage['totalAnalysis' + str(i)].calculation_again()
                if i == 0:
                    sheet = wb.active
                    sheet.title = self.analysisPage['totalAnalysis' + str(i)].ip_lineEdit.text() + '统计表'
                else:
                    sheet = wb.create_sheet(self.analysisPage['totalAnalysis' + str(i)].ip_lineEdit.text() + '统计表')
                value = [[self.analysisPage['totalAnalysis' + str(i)].label_7.text(),
                          self.analysisPage['totalAnalysis' + str(i)].label_9.text(),
                          self.analysisPage['totalAnalysis' + str(i)].label_13.text(),
                          self.analysisPage['totalAnalysis' + str(i)].label_11.text(),
                          self.analysisPage['totalAnalysis' + str(i)].label_16.text()],
                         [self.analysisPage['totalAnalysis' + str(i)].cpu_label.text(),
                          self.analysisPage['totalAnalysis' + str(i)].men_label.text(),
                          self.analysisPage['totalAnalysis' + str(i)].disk_label.text(),
                          self.analysisPage['totalAnalysis' + str(i)].net_label.text(),
                          self.analysisPage['totalAnalysis' + str(i)].IOPS_label.text()]]
                for i in range(0, 2):
                    for j in range(0, len(value[i])):
                        sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
            filePath, ok2 = QFileDialog.getSaveFileName(self,
                                                        "文件保存",
                                                        "temp/",
                                                        "Xlsx Files (*.xlsx);;All Files (*)")
            if filePath == '':
                return
            wb.save(filePath)
            self.Tips('写入数据成功！')
            self.analysis_pushButton.setDisabled(False)
        except Exception as e:
            self.Tips('导出execl异常，请重试')

    # 提示窗口
    def Tips(self, message):
        QMessageBox.about(self, "提示", message)


# 密码页面
class password_form(QDialog, Ui_password_Dialog):
    # 初始化
    def __init__(self, parent=None):
        super(password_form, self).__init__(parent)
        self.setupUi(self)
        self.unlock_pushButton.clicked.connect(self.unlock)

    def unlock(self):
        conf = config()
        if self.password_lineEdit.text() == conf.decrypt(conf.getOption('sysconfig', 'password')):
            WindowShow.show()
            self.close()
        else:
            self.Tips("密码错误,请重试")
            self.widget.setStyleSheet("border-image: url(:/icon/error_password.jpg);")

    # 提示窗口
    def Tips(self, message):
        QMessageBox.about(self, "提示", message)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    win = password_form()
    WindowShow = MyMainWindow()
    win.show()
    sys.exit(app.exec_())
